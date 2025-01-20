from flask import Flask, render_template, request, redirect, url_for, jsonify, session
import openpyxl
from datetime import datetime
import os
import secrets 

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # Set the secret key for session management

# Define the path to the main workbook
LOG_WORKBOOK_PATH = 'user_log.xlsx'

# Function to log user actions (login/logout) to different sheets
def log_user_action(username, user_type, timestamp, action, description=""):
    # Determine which sheet to use based on user type and action (login/logout)
    if user_type.lower() == 'supplier':
        sheet_name = 'Login Supplier' if action.lower() == 'login' else 'Logout Supplier'
    elif user_type.lower() == 'installer':
        sheet_name = 'Login Installer' if action.lower() == 'login' else 'Logout Installer'
    else:
        return

    try:
        # Create the workbook if it doesn't exist, or open it if it exists
        if not os.path.exists(LOG_WORKBOOK_PATH):
            workbook = openpyxl.Workbook()
            workbook.create_sheet('Login Supplier')
            workbook.create_sheet('Logout Supplier')
            workbook.create_sheet('Login Installer')
            workbook.create_sheet('Logout Installer')
            workbook.save(LOG_WORKBOOK_PATH)
        else:
            workbook = openpyxl.load_workbook(LOG_WORKBOOK_PATH)

        # Access the appropriate sheet
        sheet = workbook[sheet_name]

        # If the sheet is empty (first time logging data), add the headers
        if sheet.max_row == 1:
            sheet.append(['Username', 'UserType', 'Timestamp', 'Action', 'Description'])

        # Log the user action (append new data to the sheet)
        sheet.append([username, user_type, timestamp, action, description])
        workbook.save(LOG_WORKBOOK_PATH)
    except Exception as e:
        print(f"Error writing to {LOG_WORKBOOK_PATH}: {e}")

# Function to get usernames based on user type (supplier or installer)
def get_usernames(user_type):
    workbook_path = 'suppliers.xlsx' if user_type.lower() == 'supplier' else 'installers.xlsx'
    sheet_name = 'supplier' if user_type.lower() == 'supplier' else 'installer'

    try:
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook[sheet_name]
        usernames = [row[0].value for row in sheet.iter_rows(min_row=2)]
        return usernames
    except Exception as e:
        print(f"Error reading {workbook_path}: {e}")
        return []

# Route to display login page
@app.route('/')
def login():
    return render_template('login.html')

# API route to get a list of usernames based on user type (supplier/installer)
@app.route('/get_usernames')
def get_usernames_api():
    user_type = request.args.get('type')
    usernames = get_usernames(user_type)
    return jsonify({'usernames': usernames})

# Route to handle login and show a welcome page
@app.route('/welcome', methods=['POST'])
def welcome():
    user_type = request.form['user_type']
    username = request.form['username']
    login_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    session['username'] = username
    session['user_type'] = user_type
    session['login_time'] = login_time

    # Log the user login action
    log_user_action(username, user_type, login_time, action="login")

    return render_template('welcome.html', username=username, user_type=user_type, login_time=login_time)

# Route to handle logout and log the logout action
@app.route('/logout', methods=['POST'])
def logout():
    # Clear the session and redirect to login
    session.clear()
    return redirect(url_for('login'))

# Main entry point to run the Flask application
if __name__ == '__main__':
    app.run(debug=True)
