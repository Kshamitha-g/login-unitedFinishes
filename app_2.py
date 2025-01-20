from flask import Flask, render_template, request, redirect, url_for, jsonify,session
import openpyxl
from datetime import datetime
import os
import secrets 

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
# Define the path to the main workbook
LOG_WORKBOOK_PATH = 'user_log.xlsx'

# Function to log user actions (login/logout) to different sheets
def log_user_action(username, user_type, timestamp, action, description=""):
    # Determine which sheet to use based on user type and action (login/logout)
    if user_type.lower() == 'supplier':
        if action.lower() == 'login':
            sheet_name = 'Login Supplier'
        elif action.lower() == 'logout':
            sheet_name = 'Logout Supplier'
        else:
            return
    elif user_type.lower() == 'installer':
        if action.lower() == 'login':
            sheet_name = 'Login Installer'
        elif action.lower() == 'logout':
            sheet_name = 'Logout Installer'
        else:
            return
    else:
        return

    try:
        # Create the workbook if it doesn't exist, or open it if it exists
        if not os.path.exists(LOG_WORKBOOK_PATH):
            workbook = openpyxl.Workbook()
            workbook.create_sheet('Login Supplier')
            workbook.create_sheet('Logout Supplier')
            workbook.create_sheet('Login Installer')
            workbook.create_sheet('Logout Installer')
            workbook.save(LOG_WORKBOOK_PATH)
        else:
            workbook = openpyxl.load_workbook(LOG_WORKBOOK_PATH)

        # Access the appropriate sheet
        sheet = workbook[sheet_name]

        # If the sheet is empty (first time logging data), add the headers
        if sheet.max_row == 1:
            sheet.append(['Username', 'UserType', 'Timestamp', 'Action', 'Description'])

        # Log the user action (append new data to the sheet)
        sheet.append([username, user_type, timestamp, action, description])
        workbook.save(LOG_WORKBOOK_PATH)
    except Exception as e:
        print(f"Error writing to {LOG_WORKBOOK_PATH}: {e}")

# Function to get usernames based on user type (supplier or installer)
def get_usernames(user_type):
    if user_type.lower() == 'supplier':
        workbook_path = 'suppliers.xlsx'
        sheet_name = 'supplier'
    elif user_type.lower() == 'installer':
        workbook_path = 'installers.xlsx'
        sheet_name = 'installer'
    else:
        return []

    try:
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook[sheet_name]
        usernames = [row[0].value for row in sheet.iter_rows(min_row=2)]
        return usernames
    except Exception as e:
        print(f"Error reading {workbook_path}: {e}")
        return []

# Route to display login page
@app.route('/')
def login():
    return render_template('login.html')

# API route to get a list of usernames based on user type (supplier/installer)
@app.route('/get_usernames')
def get_usernames_api():
    user_type = request.args.get('type')
    usernames = get_usernames(user_type)
    return jsonify({'usernames': usernames})

# Route to handle login and show a welcome page
'''@app.route('/welcome', methods=['POST'])
def welcome():
    user_type = request.form['user_type']
    username = request.form['username']
    login_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Log the user login action
    log_user_action(username, user_type, login_time, action="login")
    
    return render_template('welcome.html', username=username, user_type=user_type, login_time=login_time)'''
'''@app.route('/welcome', methods=['POST'])
def welcome():
    # Safely retrieve 'username' and 'user_type' using .get() to avoid KeyError
    user_type = request.form.get('user_type')
    username = request.form.get('username')
    
    if not username or not user_type:
        return "Missing username or user type", 400

    login_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_user_action(username, user_type, login_time, action="login")
    
    return render_template('welcome.html', username=username, user_type=user_type, login_time=login_time)'''

@app.route('/welcome', methods=['POST'])
def welcome():
    user_type = request.form['user_type']
    username = request.form['username']
    login_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    session['username'] = username
    session['user_type'] = user_type
    session['login_time'] = login_time

    # Save login details to Excel (you can implement this as needed)

    return render_template('welcome.html', username=username, user_type=user_type, login_time=login_time)



# Route to handle logout and log the logout action
''''@app.route('/logout', methods=['POST'])
def logout():
    description = request.form['description']
    username = request.form['username']
    user_type = request.form['user_type']
    logout_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Log the user logout action
    log_user_action(username, user_type, logout_time, action="logout", description=description)
    
    return redirect(url_for('login'))'''
'''@app.route('/logout', methods=['POST'])
def logout():
    # Use .get() to avoid KeyError
    username = request.form.get('username')
    user_type = request.form.get('user_type')
    description = request.form.get('description', '')  # Optional field

    # Check if the necessary data is present
    if not username or not user_type:
        return "Missing username or user type", 400

    logout_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_user_action(username, user_type, logout_time, action="logout", description=description)

    return redirect(url_for('login'))'''
'''def logout():
    session.clear()  # Clears session data
    return redirect(url_for('login'))'''
@app.route('/logout', methods=['POST'])
def logout():
    # Clear the session and redirect to login
    session.clear()
    return redirect(url_for('login'))



# Main entry point to run the Flask application
if __name__ == '__main__':
    app.run(debug=True)
