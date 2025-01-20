from flask import Flask, render_template, request, redirect, url_for, jsonify
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)

# Define paths to Excel workbooks
SUPPLIER_DATA_PATH = 'suppliers.xlsx'
INSTALLER_DATA_PATH = 'installers.xlsx'
SUPPLIER_LOG_PATH = 'supplier_log.xlsx'
INSTALLER_LOG_PATH = 'installer_log.xlsx'

def get_usernames(user_type):
    if user_type.lower() == 'supplier':
        workbook_path = SUPPLIER_DATA_PATH
        sheet_name = 'supplier'
    elif user_type.lower() == 'installer':
        workbook_path = INSTALLER_DATA_PATH
        sheet_name = 'installer'
    else:
        return []

    # Load workbook and read usernames
    try:
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook[sheet_name]
        usernames = [row[0].value for row in sheet.iter_rows(min_row=2)]
        return usernames
    except Exception as e:
        print(f"Error reading {workbook_path}: {e}")
        return []

@app.route('/')
def login():
    return render_template('login.html')

@app.route('/get_usernames')
def get_usernames_api():
    user_type = request.args.get('type')
    usernames = get_usernames(user_type)
    return jsonify({'usernames': usernames})

@app.route('/welcome', methods=['POST'])
def welcome():
    user_type = request.form['user_type']
    username = request.form['username']
    login_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    log_user_action(username, user_type, login_time, action="login")
    
    return render_template('welcome.html', username=username, user_type=user_type, login_time=login_time)

@app.route('/logout', methods=['POST'])
def logout():
    description = request.form['description']
    username = request.form['username']
    user_type = request.form['user_type']
    logout_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    log_user_action(username, user_type, logout_time, action="logout", description=description)
    
    return redirect(url_for('login'))

def log_user_action(username, user_type, timestamp, action, description=""):
    if user_type.lower() == 'supplier':
        workbook_path = SUPPLIER_LOG_PATH
        sheet_name = 'SupplierLog'
    elif user_type.lower() == 'installer':
        workbook_path = INSTALLER_LOG_PATH
        sheet_name = 'InstallerLog'
    else:
        return

    try:
        if not os.path.exists(workbook_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            sheet.append(['Username', 'UserType', 'Timestamp', 'Action', 'Description'])
        else:
            workbook = openpyxl.load_workbook(workbook_path)
            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(['Username', 'UserType', 'Timestamp', 'Action', 'Description'])
            else:
                sheet = workbook[sheet_name]

        sheet.append([username, user_type, timestamp, action, description])
        workbook.save(workbook_path)
    except Exception as e:
        print(f"Error writing to {workbook_path}: {e}")

if __name__ == '__main__':
    app.run(debug=True)
